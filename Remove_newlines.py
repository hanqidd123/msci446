import openpyxl
import nltk

wb = openpyxl.load_workbook('clean_data.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

for x in range(1, 4481):
    print(x)
    if sheet["C" + str(x)].value is None:
        sheet.cell(row=x, column=4).value = "blank job description"
    else:
        sentence = sheet["C" + str(x)].value.replace('\n', ' ')
        print(sentence)
        sheet.cell(row=x, column=4).value = sentence
        print(sheet.cell(row=x, column=4).value)

wb.save('clean_data.xlsx')


