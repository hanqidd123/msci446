import openpyxl
import nltk
from nltk.corpus import stopwords
from nltk.stem import SnowballStemmer

from nltk.tokenize import RegexpTokenizer
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer

stop_words = set(stopwords.words("english"))
wb = openpyxl.load_workbook('clean_data.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
dataStr = wb.get_sheet_by_name("data")
matrix = wb.get_sheet_by_name("matrix")
ps = SnowballStemmer("english")
regex_token = RegexpTokenizer(r'\w+')
unique_word_list = []
discipline_objective = {}

#ps = SnowballStemmer("english")
regex_token = RegexpTokenizer(r'\w+')
unique_word_list = []
original_word_list = {}
#4481

for x in range(1, 4481):
    print(x)
    sentence = sheet["D" + str(x)].value

    tokens = regex_token.tokenize(sentence)
    #sentences with no stop words
    filtered_sentence_stopwords = []
    for w in tokens:
        if w not in stop_words:
            w = w.lower()
            filtered_sentence_stopwords.append(w)

    current_discipline = {}
    #sentence after stemming
    filtered_sentence_stemmed = []
    for stem_words in filtered_sentence_stopwords:
        if ps.stem(stem_words) not in original_word_list:
            value_list = []
            value_list.append(stem_words)
            original_word_list[ps.stem(stem_words)] = value_list

        else:
            original_word_list[ps.stem(stem_words)] = list(set(original_word_list[ps.stem(stem_words)]))
            original_word_list[ps.stem(stem_words)].append(stem_words)


        filtered_sentence_stemmed.append(ps.stem(stem_words))
        unique_word_list.append(ps.stem(stem_words))


    #tagged = nltk.pos_tag(filtered_sentence_stopwords)



    unique_word_list.append(w)
    filtered_sentence_stopwords.sort()

    #sentence after stemming
    #filtered_sentence_stemmed = []
    #for stem_words in filtered_sentence_stopwords:
       # filtered_sentence_stemmed.append(ps.stem(stem_words))
        #unique_word_list.append(ps.stem(stem_words))

    #filtered_sentence_stemmed = list(set(filtered_sentence_stemmed))
    #filtered_sentence_stemmed.sort()

    #tagged = nltk.pos_tag(filtered_sentence_stopwords)
    #split_string = str(tagged)

    dataStr.cell(row=1, column=x).value = sheet["B" + str(x)].value
    dataStr.cell(row=2, column=x).value = sheet["A" + str(x)].value.replace(" ","")
    dataStr.cell(row=3, column=x).value = " ".join(filtered_sentence_stemmed)
    print(dataStr.cell(row=2, column=x).value)



#finding unique words
print(original_word_list)

wb.save('clean_data_test.xlsx')






