import openpyxl
import nltk
from nltk.corpus import stopwords
from nltk.stem.lancaster import LancasterStemmer
stop_words = set(stopwords.words("english"))
wb = openpyxl.load_workbook('clean_data.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
dataStr = wb.get_sheet_by_name("data")
ps = LancasterStemmer()
for x in range(1, 4481):
    print(x)
    sentence = sheet["D" + str(x)].value
    tokens = nltk.word_tokenize(sentence)
    print(tokens)
    #sentences with no stop words
    filtered_sentence_stopwords = []

    for w in tokens:
        if w not in stop_words:
            filtered_sentence_stopwords.append(w)
    #sentence after stemming
    #filtered_sentence_stemmed = []
    #for stem_words in filtered_sentence_stopwords:
        #filtered_sentence_stemmed.append(ps.stem(stem_words))
    #print(filtered_sentence_stemmed)

    tagged = nltk.pos_tag(filtered_sentence_stopwords)
    print(tagged)
    #split_string = str(tagged)
    count = 3
    length = len(tagged)
    dataStr.cell(row=1, column=x).value = sheet["B" + str(x)].value
    dataStr.cell(row=2, column=x).value = sheet["A" + str(x)].value
    for item in tagged:
        dataStr.cell(row=count, column=x).value = str(item)
        print(dataStr.cell(row=count, column=x).value)
        count = count + 1

wb.save('clean_data.xlsx')






