from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.metrics import adjusted_rand_score
import matplotlib.pyplot as plt
import openpyxl
import numpy as np

wb = openpyxl.load_workbook('ENG.xlsx')

#matrix2 = wb.get_sheet_by_name("matrix")

dataStr = wb.get_sheet_by_name("mgte")
open_sheet = wb.get_sheet_by_name("result")
documents = []
count = int(dataStr.cell(row=3, column=1).value)
print(count)

for i in range(1,count+1):
    documents.append(dataStr.cell(row=2, column=i).value)

#for j in range(1,28):
true_k = 35
vectorizer = TfidfVectorizer()
x = vectorizer.fit_transform(documents)
#print(x)
model = KMeans(n_clusters=true_k, init="k-means++", max_iter=100, n_init=1)
model.fit(x)
order_centroids = model.cluster_centers_.argsort()[:,::-1]
terms = vectorizer.get_feature_names()
#print("The cluster centroids are: \n", model.cluster_centers_)
#print("Cluster", model.labels_)
#print("Sum of distances of samples to their closest cluster center: ", model.inertia_)
open_sheet.cell(row=i + 1, column=2).value = model.inertia_
print(open_sheet.cell(row=i + 1, column=2).value)
for i in range(true_k):
        print(" ")
        print("Cluster" + str(i))
        for ind in order_centroids[i, :10]:
            print(terms[ind])
#wb.save('ENG_result.xlsx')




    #average = total_inertia / true_k
    #print("Average Inertia for ", true_k, "clusters is" , average)
