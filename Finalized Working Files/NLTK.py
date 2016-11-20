import openpyxl
import nltk
from nltk.corpus import stopwords
from nltk.stem import SnowballStemmer

#from nltk.stem.snowball import SnowballStemmer
#from nltk.stem.lancaster import LancasterStemmer
from nltk.tokenize import RegexpTokenizer
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer

stop_words = set(stopwords.words("english"))
wb = openpyxl.load_workbook('input_removeBlankDesp.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
dataStr = wb.get_sheet_by_name("data")
matrix = wb.get_sheet_by_name("matrix")
ps = SnowballStemmer("english")
regex_token = RegexpTokenizer(r'\w+')
unique_word_list = []
discipline_objective = {}

#ps = SnowballStemmer("english")
regex_token = RegexpTokenizer(r'\w+')
unique_word_list = []
original_word_list = {}
#4481

for x in range(1, 4470):
    print(x)
    sentence = sheet["D" + str(x)].value

    tokens = regex_token.tokenize(sentence)
    #sentences with no stop words
    filtered_sentence_stopwords = []
    for w in tokens:
        if w not in stop_words:
            w = w.lower()
            filtered_sentence_stopwords.append(w)

    current_discipline = {}
    #sentence after stemming
    filtered_sentence_stemmed = []
    for stem_words in filtered_sentence_stopwords:
        if ps.stem(stem_words) not in original_word_list:
            value_list = []
            value_list.append(stem_words)
            original_word_list[ps.stem(stem_words)] = value_list

        else:
            original_word_list[ps.stem(stem_words)] = list(set(original_word_list[ps.stem(stem_words)]))
            original_word_list[ps.stem(stem_words)].append(stem_words)


        filtered_sentence_stemmed.append(ps.stem(stem_words))
        unique_word_list.append(ps.stem(stem_words))


    #tagged = nltk.pos_tag(filtered_sentence_stopwords)
    split_string = str(filtered_sentence_stemmed)

    unique_word_list.append(w)
    filtered_sentence_stopwords.sort()

    #sentence after stemming
    #filtered_sentence_stemmed = []
    #for stem_words in filtered_sentence_stopwords:
       # filtered_sentence_stemmed.append(ps.stem(stem_words))
        #unique_word_list.append(ps.stem(stem_words))

    #filtered_sentence_stemmed = list(set(filtered_sentence_stemmed))
    #filtered_sentence_stemmed.sort()

    tagged = nltk.pos_tag(filtered_sentence_stopwords)
    split_string = str(tagged)

    count = 3
    length = len(filtered_sentence_stemmed)
    dataStr.cell(row=1, column=x).value = sheet["B" + str(x)].value
    dataStr.cell(row=2, column=x).value = sheet["A" + str(x)].value
    for item in filtered_sentence_stemmed:
        dataStr.cell(row=count, column=x).value = str(item)
        count = count + 1


# finding unique words
print(original_word_list)
unique_word_list = list(set(unique_word_list))
unique_word_list.sort()
count_x = 2
print(unique_word_list)

for item in unique_word_list:
    matrix.cell(row=1, column=count_x).value = str(item)
    count_x = count_x + 1

discipline_list = []

for x in range(1,4470):
    discipline_list.append(dataStr.cell(row=2, column=x).value)

discipline_list = list(set(discipline_list))

count_y = 2
for item in discipline_list:
    matrix.cell(row=count_y, column=1).value = str(item)
    count_y = count_y + 1

wb.save ("testingOutput-2.xlsx")






