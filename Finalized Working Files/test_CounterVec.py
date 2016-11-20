import openpyxl
from sklearn.feature_extraction.text import CountVectorizer

wb = openpyxl.load_workbook('countVectorizer_input_1.xlsx')
inputSheet = wb.get_sheet_by_name("Sheet2")
outputSheet = wb.get_sheet_by_name("matrix")
dest_fileName = "testOutput-1.xlsx"

tag = []

for x in range(1, 4475):
    uniDescription = inputSheet["C" + str(x)].value
    tag.append(uniDescription)

vec_count = CountVectorizer()
data_countVec = vec_count.fit_transform(tag).toarray()
vocab_countVec = vec_count.get_feature_names()

wordColumn = 3
for uniWord in vocab_countVec:
    outputSheet.cell(row=1, column=wordColumn).value = str(uniWord)
    wordColumn = wordColumn + 1

dataRow = 2
for jobDescription in data_countVec:
    dataColumn = 3
    for wordInJob in jobDescription:
        outputSheet.cell(row=dataRow, column=dataColumn).value = int(wordInJob)
        dataColumn = dataColumn + 1
    dataRow = dataRow + 1

print(vocab_countVec)
print(data_countVec)

wb.save(filename=dest_fileName)