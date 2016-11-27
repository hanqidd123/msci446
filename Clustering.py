from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.metrics import adjusted_rand_score
import matplotlib.pyplot as plt
import openpyxl
import numpy as np

wb = openpyxl.load_workbook('clean_data_test.xlsx')
dataStr = wb.get_sheet_by_name("data")
matrix2 = wb.get_sheet_by_name("matrix")
documents = []
count = int(dataStr.cell(row=4, column=1).value)
print(count)

for i in range(1,count+1):
    documents.append(dataStr.cell(row=2, column=i).value)


true_k = 90
vectorizer = TfidfVectorizer()
x = vectorizer.fit_transform(documents)

print(x.shape)
model = KMeans(n_clusters=true_k, init="k-means++", random_state=170, max_iter=100, n_init=1)

model.fit(x)
clusters = model.labels_.tolist()
from sklearn.externals import joblib




order_centroids = model.cluster_centers_.argsort()[:,::-1]
print(len(order_centroids))
print(order_centroids)
terms = vectorizer.get_feature_names()
print("The cluster centroids are: \n", model.cluster_centers_)
print("Cluster", model.labels_)
clusters = model.labels_
print()
print("Sum of distances of samples to their closest cluster center: ", model.inertia_)
#matrix2.cell(row=i + 1, column=2).value = model.inertia_
for i in range(true_k):
    print(" ")
    print("Cluster" + str(i))
    print(len(order_centroids[i]))
    for ind in order_centroids[i, :8]:
        print(terms[ind])

clustering = open('clusters.txt','w')

for entry in clusters:


    clustering.write(str(entry) + '\n')
clustering.close()


    #average = total_inertia / true_k
    #print("Average Inertia for ", true_k, "clusters is" , average)
#wb.save('ENG.xlsx')